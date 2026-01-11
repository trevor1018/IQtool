import os
import cv2
import openpyxl
import numpy as np
import time
import matplotlib.pyplot as plt
import matplotlib.patches as patches
import re
import tkinter as tk
from tkinter import filedialog
from PIL import Image

def atoi(text):
    return int(text) if text.isdigit() else text

def natural_keys(text):
    return [ atoi(c) for c in re.split(r'(\d+)', text) ]

def file_filter(f):
    if f[-5:] in ['.exif'] or f[-4:] in ['.txt']:
        return True
    else:
        return False

def file_filter_jpg(f):
    if f[-4:] in ['.jpg', '.JPG']:
        return True
    else:
        return False

def create_xls(file_path):
    fn = 'mtkAWBanalysis.xlsm'
    wb = openpyxl.load_workbook(fn, read_only=False, keep_vba=True)
    wb.active = 0
    ws = wb.active
    ws.cell(column=2, row=6).value = file_path.split("/")[-3]
    
    f = open(file_path, "r")
    
    wb.active = 1
    ws = wb.active
    P1 = []
    for line in f:
        if "AWB Light source probability" in line:
            line2, line3, line4, line5, line6, line7, line8, line9, line10, line11 = [next(f) for _ in range(10)]
            P1.append(re.sub("[^0-9-,]","", line5).split(","))
            P1.append(re.sub("[^0-9-,]","", line6).split(","))
            P1.append(re.sub("[^0-9-,]","", line7).split(","))
            P1.append(re.sub("[^0-9-,]","", line8).split(","))
            P1.append(re.sub("[^0-9-,]","", line11).split(","))
            P1.append(re.sub("[^0-9-,]","", line9).split(","))
            P1.append(re.sub("[^0-9-,]","", line10).split(","))
        
    for i in range(0,7):
        for j in range(0,23):
            ws.cell(column=3+i, row=35+j).value = int(P1[i][j])
    print("AWB.cpp is ok!")
    return wb

print("mtkAWBanalysis is runing...")

root = tk.Tk()
root.withdraw()
file_path = filedialog.askopenfilename()
print(file_path)

refer = input("Have reference or not (0: no, 1: yes): ")
refer = int(refer)

localtime = time.localtime()
clock = str(60*60*localtime[3] + 60*localtime[4] + localtime[5])

yourPath = "Exif"
allFileList = os.listdir(yourPath)
allFileList_exif = np.sort(allFileList,axis=0)
allFileList_exif = list(filter(file_filter, allFileList_exif))
allFileList_exif.sort(key=natural_keys)
allFileList_jpg = np.sort(allFileList,axis=0)
allFileList_jpg = list(filter(file_filter_jpg, allFileList_jpg))
allFileList_jpg.sort(key=natural_keys)

for i in range(0,(np.size(allFileList_exif))):
    path_name = yourPath + "/" + allFileList_exif[i]
    exifFile = open(path_name, "r")
    file_name = os.path.basename(path_name)
    base = os.path.splitext(file_name)[0]
    baseTag = base.split(".")[0]
    
    if i % 20 == 0:
        startNum = re.sub("[^0-9-,]","", base[0:2])
        wb = create_xls(file_path)
    
    sheet = wb[wb.sheetnames[0]]
    target = wb.copy_worksheet(sheet)
    target.title = baseTag
    wb.active = int((i%20)+2)
    ws = wb.active
    
    print(base)
    
    CCT = []
    AWB_TAG_NEUTRAL_PB_NUM_D = []
    AWB_TAG_P1_D = []
    AWB_TAG_P2_D = []
    AWB_TAG_STA_GAIN_R_D = []
    AWB_TAG_STA_GAIN_B_D = []
    AWB_TAG_SPAT_GAIN_R_D = []
    AWB_TAG_SPAT_GAIN_B_D = []
    AWB_TAG_DAYLIGHT_PROB_D = []
    AWB_TAG_AVG_YR_D = []
    AWB_TAG_STAT_LIMIT_Y_D = []
    AWB_TAG_LIMIT_D = []
    AWB_TAG_STAT_LIMIT_W_RED_D = []
    AWB_TAG_STAT_LIMIT_PROJ_W_D = []
    AWB_TAG_PB_NUM_THR_D = []
    AWB_TAG_SPAT_LV_THR_L_D = []
    AWB_TAG_SPAT_LV_THR_H_D = []
    # AWB_TAG_HIT_NR_D = []
    
    for t in range(0,8):
        locals()['exclude_'+str(t)] = []
        
    for t in range(0,8):
        locals()['extra_'+str(t)] = []
        
    locals()['AWB_TAG_EXTRACOLOR_ENABLE'] = []
    
    T_area = []
    WF_area = []
    F_area = []
    CWF_area = []
    D_area = []
    S_area = []
    DF_area = []
    
    Sub_F_area = []
    Sub_CWF_area = []
    
    for line in exifFile:
        if "AWB_TAG_ALGO_SCENE_LV" in line:
            ws.cell(column=10, row=3).value = int(re.sub("[^0-9-,]","", line))
        if "AWB_TAG_CCT" in line:
            CCT.append(re.sub("[^0-9-,]","", line))
        
        if "AWB_TAG_OUTPUT_CAL_GAIN_R" in line:
            ws.cell(column=8, row=7).value = int(re.sub("[^0-9-,]","", line))
        if "AWB_TAG_OUTPUT_CAL_GAIN_B" in line:
            ws.cell(column=10, row=7).value = int(re.sub("[^0-9-,]","", line))
        
        if "AWB_TAG_NONEUTRAL_EQV_GAIN_R" in line:
            ws.cell(column=5, row=15).value = int(re.sub("[^0-9-,]","", line))
        if "AWB_TAG_NONEUTRAL_EQV_GAIN_B" in line:
            ws.cell(column=5, row=16).value = int(re.sub("[^0-9-,]","", line))
        if "AWB_TAG_RELIABLE_MODE" in line:
            ws.cell(column=10, row=12).value = int(re.sub("[^0-9-,]","", line))
        if "AWB_NVRAM_TEMPORAL_ENQ_NEUTRAL_BLK_THR" in line:
            ws.cell(column=15, row=13).value = int(re.sub("[^0-9-,]","", line))
        if "AWB_NVRAM_TEMPORAL_ENQ_CWF_DF_BLK_THR" in line:
            ws.cell(column=15, row=14).value = int(re.sub("[^0-9-,]","", line))
        if "AWB_TAG_PB_NUM_THR_NONNEUTRAL" in line:
            ws.cell(column=15, row=16).value = int(re.sub("[^0-9-,]","", line))
        if "AWB_NVRAM_PREF_COLOR_OFFSET_THR_T" in line:
            ws.cell(column=24, row=15).value = int(re.sub("[^0-9-,]","", line))
        if "AWB_NVRAM_PREF_COLOR_OFFSET_THR_WF" in line:
            ws.cell(column=24, row=16).value = int(re.sub("[^0-9-,]","", line))
        if "AWB_NVRAM_PREF_COLOR_OFFSET_THR_S" in line:
            ws.cell(column=24, row=17).value = int(re.sub("[^0-9-,]","", line))
        
        if "AWB_TAG_NEUTRAL_PB_NUM_T" in line:
            ws.cell(column=5, row=20).value = int(re.sub("[^0-9-,]","", line))
        if "AWB_TAG_NEUTRAL_PB_NUM_WF" in line:
            ws.cell(column=5, row=21).value = int(re.sub("[^0-9-,]","", line))
        if "AWB_TAG_NEUTRAL_PB_NUM_F" in line:
            ws.cell(column=5, row=22).value = int(re.sub("[^0-9-,]","", line))
        if "AWB_TAG_NEUTRAL_PB_NUM_CWF" in line:
            ws.cell(column=5, row=23).value = int(re.sub("[^0-9-,]","", line))
        if "AWB_TAG_NEUTRAL_PB_NUM_D" in line:
            AWB_TAG_NEUTRAL_PB_NUM_D.append(re.sub("[^0-9-,]","", line))
        if "AWB_TAG_NEUTRAL_PB_NUM_S" in line:
            ws.cell(column=5, row=25).value = int(re.sub("[^0-9-,]","", line))
        if "AWB_TAG_NEUTRAL_PB_NUM_DF" in line:
            ws.cell(column=5, row=26).value = int(re.sub("[^0-9-,]","", line))
        
        if "AWB_TAG_P1_T" in line:
            ws.cell(column=5, row=57).value = int(re.sub("[^0-9-,]","", line)[1:])
            ws.cell(column=1, row=57).value = int(re.sub("[^0-9-,]","", line)[1:])
        if "AWB_TAG_P1_WF" in line:
            ws.cell(column=5, row=58).value = int(re.sub("[^0-9-,]","", line)[1:])
            ws.cell(column=1, row=58).value = int(re.sub("[^0-9-,]","", line)[1:])
        if "AWB_TAG_P1_F" in line:
            ws.cell(column=5, row=59).value = int(re.sub("[^0-9-,]","", line)[1:])
            ws.cell(column=1, row=59).value = int(re.sub("[^0-9-,]","", line)[1:])
        if "AWB_TAG_P1_CWF" in line:
            ws.cell(column=5, row=60).value = int(re.sub("[^0-9-,]","", line)[1:])
            ws.cell(column=1, row=60).value = int(re.sub("[^0-9-,]","", line)[1:])
        if "AWB_TAG_P1_D" in line:
            AWB_TAG_P1_D.append(re.sub("[^0-9-,]","", line)[1:])
        if "AWB_TAG_P1_S" in line:
            ws.cell(column=5, row=62).value = int(re.sub("[^0-9-,]","", line)[1:])
            ws.cell(column=1, row=62).value = int(re.sub("[^0-9-,]","", line)[1:])
        if "AWB_TAG_P1_DF" in line:
            ws.cell(column=5, row=63).value = int(re.sub("[^0-9-,]","", line)[1:])
            ws.cell(column=1, row=63).value = int(re.sub("[^0-9-,]","", line)[1:])
            
        if "AWB_TAG_P2_F" in line:
            ws.cell(column=10, row=59).value = int(re.sub("[^0-9-,]","", line)[1:])
        if "AWB_TAG_P2_CWF" in line:
            ws.cell(column=10, row=60).value = int(re.sub("[^0-9-,]","", line)[1:])
        if "AWB_TAG_P2_D" in line:
            AWB_TAG_P2_D.append(re.sub("[^0-9-,]","", line)[1:])
        if "AWB_TAG_P2_DF" in line:
            ws.cell(column=10, row=63).value = int(re.sub("[^0-9-,]","", line)[1:])
            
        if "AWB_TAG_IS_ABOVE_DAY_LOCUS_T" in line:
            ws.cell(column=14, row=58).value = int(re.sub("[^0-9-,]","", line))
        if "AWB_TAG_GM_OFFSET_T" in line:
            ws.cell(column=14, row=59).value = int(re.sub("[^0-9-,]","", line))
        if "AWB_TAG_WEIGHT_T" in line:
            ws.cell(column=16, row=59).value = int(re.sub("[^0-9-,]","", line))
            ws.cell(column=17, row=59).value = int(re.sub("[^0-9-,]","", line))
        if "AWB_TAG_IS_ABOVE_DAY_LOCUS_WF" in line:
            ws.cell(column=14, row=60).value = int(re.sub("[^0-9-,]","", line))
        if "AWB_TAG_GM_OFFSET_WF" in line:
            ws.cell(column=14, row=61).value = int(re.sub("[^0-9-,]","", line))
        if "AWB_TAG_WEIGHT_WF" in line:
            ws.cell(column=16, row=61).value = int(re.sub("[^0-9-,]","", line))
            ws.cell(column=17, row=61).value = int(re.sub("[^0-9-,]","", line))
        if "AWB_TAG_IS_ABOVE_DAY_LOCUS_S" in line:
            ws.cell(column=14, row=62).value = int(re.sub("[^0-9-,]","", line))
        if "AWB_TAG_GM_OFFSET_S" in line:
            ws.cell(column=14, row=63).value = int(re.sub("[^0-9-,]","", line))
        if "AWB_TAG_WEIGHT_S" in line:
            ws.cell(column=16, row=63).value = int(re.sub("[^0-9-,]","", line))
            ws.cell(column=17, row=63).value = int(re.sub("[^0-9-,]","", line))
        
        for k in range(0,8):
            if "AWB_TAG_EXCLUDECOLOR_"+str(k)+"_INFO_COUNT" in line:
                ws.cell(column=9+(k+1)*5, row=22).value = int(re.sub("[^0-9-,]","", line)[1:])
                ws.cell(column=9+(k+1)*5, row=28).value = int(re.sub("[^0-9-,]","", line)[1:])
            if "AWB_TAG_EXCLUDE_"+str(k)+"_WEIGHT" in line:
                ws.cell(column=11+(k+1)*5, row=22).value = int(re.sub("[^0-9-,]","", line)[1:])
                ws.cell(column=12+(k+1)*5, row=22).value = int(re.sub("[^0-9-,]","", line)[1:])
            if "AWB_TAG_EXCLUDE_"+str(k)+"_G_RANGE" in line:
                ws.cell(column=9+(k+1)*5, row=23).value = int(re.sub("[^0-9-,]","", line)[1:])
            if "AWB_TAG_EXCLUDE_"+str(k)+"_LV_RANGE" in line:
                ws.cell(column=11+(k+1)*5, row=23).value = int(re.sub("[^0-9-,]","", line)[1:])
                ws.cell(column=12+(k+1)*5, row=23).value = int(re.sub("[^0-9-,]","", line)[1:])
            if "AWB_TAG_EXCLUDE_"+str(k)+"_AREA_U" in line:
                locals()['exclude_'+str(k)].append(re.sub("[^0-9-,]","", line)[1:])
                ws.cell(column=9+(k+1)*5, row=24).value = int(re.sub("[^0-9-,]","", line)[1:])
            if "AWB_TAG_EXCLUDE_"+str(k)+"_AREA_D" in line:
                locals()['exclude_'+str(k)].append(re.sub("[^0-9-,]","", line)[1:])
                ws.cell(column=9+(k+1)*5, row=25).value = int(re.sub("[^0-9-,]","", line)[1:])
            if "AWB_TAG_EXCLUDE_"+str(k)+"_AREA_L" in line:
                locals()['exclude_'+str(k)].append(re.sub("[^0-9-,]","", line)[1:])
                ws.cell(column=11+(k+1)*5, row=24).value = int(re.sub("[^0-9-,]","", line)[1:])
            if "AWB_TAG_EXCLUDE_"+str(k)+"_AREA_R" in line:
                locals()['exclude_'+str(k)].append(re.sub("[^0-9-,]","", line)[1:])
                ws.cell(column=11+(k+1)*5, row=25).value = int(re.sub("[^0-9-,]","", line)[1:])
            if "AWB_TAG_EXCLUDE_"+str(k)+"_GAVG_L" in line:
                ws.cell(column=9+(k+1)*5, row=26).value = int(re.sub("[^0-9-,]","", line)[1:])
            if "AWB_TAG_EXCLUDE_"+str(k)+"_GAVG_H" in line:
                ws.cell(column=9+(k+1)*5, row=27).value = int(re.sub("[^0-9-,]","", line)[1:])
            if "AWB_TAG_EXCLUDE_"+str(k)+"_LV_L" in line:
                ws.cell(column=11+(k+1)*5, row=26).value = int(re.sub("[^0-9-,]","", line)[1:])
                ws.cell(column=12+(k+1)*5, row=26).value = int(re.sub("[^0-9-,]","", line)[1:])
            if "AWB_TAG_EXCLUDE_"+str(k)+"_LV_H" in line:
                ws.cell(column=11+(k+1)*5, row=27).value = int(re.sub("[^0-9-,]","", line)[1:])
                ws.cell(column=12+(k+1)*5, row=27).value = int(re.sub("[^0-9-,]","", line)[1:])
                
        for k in range(0,8):
            if "AWB_TAG_EXTRACOLOR_"+str(k)+"_MODE_WEI_GAIN" in line:
                ws.cell(column=9+(k+1)*5, row=69).value = int(re.sub("[^0-9-,]","", line)[1:])
            if "AWB_TAG_EXTRACOLOR_"+str(k)+"_MODE_WEI_P2" in line:
                ws.cell(column=10+(k+1)*5, row=69).value = int(re.sub("[^0-9-,]","", line)[2:])
            if "AWB_TAG_EXTRACOLOR_"+str(k)+"_MODE_WEI_DL_PROB" in line:
                ws.cell(column=11+(k+1)*5, row=69).value = int(re.sub("[^0-9-,]","", line)[1:])
            if "AWB_TAG_EXTRACOLOR_"+str(k)+"_ENABLE" in line:
                locals()['AWB_TAG_EXTRACOLOR_ENABLE'].append(int(re.sub("[^0-9-,]","", line)[1:]))
                ws.cell(column=9+(k+1)*5, row=70).value = int(re.sub("[^0-9-,]","", line)[1:])
                ws.cell(column=12+(k+1)*5, row=70).value = int(re.sub("[^0-9-,]","", line)[1:])
            if "AWB_TAG_EXTRACOLOR_"+str(k)+"_AREA_U" in line:
                locals()['extra_'+str(k)].append(re.sub("[^0-9-,]","", line)[1:])
                ws.cell(column=10+(k+1)*5, row=76).value = int(re.sub("[^0-9-,]","", line)[1:])
            if "AWB_TAG_EXTRACOLOR_"+str(k)+"_AREA_D" in line:
                locals()['extra_'+str(k)].append(re.sub("[^0-9-,]","", line)[1:])
                ws.cell(column=11+(k+1)*5, row=76).value = int(re.sub("[^0-9-,]","", line)[1:])
            if "AWB_TAG_EXTRACOLOR_"+str(k)+"_AREA_L" in line:
                locals()['extra_'+str(k)].append(re.sub("[^0-9-,]","", line)[1:])
                ws.cell(column=9+(k+1)*5, row=76).value = int(re.sub("[^0-9-,]","", line)[1:])
            if "AWB_TAG_EXTRACOLOR_"+str(k)+"_AREA_R" in line:
                locals()['extra_'+str(k)].append(re.sub("[^0-9-,]","", line)[1:])
                ws.cell(column=8+(k+1)*5, row=76).value = int(re.sub("[^0-9-,]","", line)[1:])
            if "AWB_TAG_EXTRACOLOR_"+str(k)+"_INFO_COUNT" in line:
                ws.cell(column=11+(k+1)*5, row=70).value = int(re.sub("[^0-9-,]","", line)[1:])
            if "AWB_TAG_EXTRACOLOR_"+str(k)+"_CONFTHR" in line:
                ws.cell(column=9+(k+1)*5, row=71).value = int(re.sub("[^0-9-,]","", line)[1:])
                ws.cell(column=12+(k+1)*5, row=71).value = int(re.sub("[^0-9-,]","", line)[1:])
            if "AWB_TAG_EXTRACOLOR_"+str(k)+"_CONF" in line:
                ws.cell(column=11+(k+1)*5, row=71).value = int(re.sub("[^0-9-,]","", line)[1:])
            if "AWB_TAG_EXTRACOLOR_"+str(k)+"_SEL_LIGHT_SRC" in line:
                ws.cell(column=9+(k+1)*5, row=72).value = int(re.sub("[^0-9-,]","", line)[1:])
            if "AWB_TAG_EXTRACOLOR_"+str(k)+"_LV_RANGE" in line:
                ws.cell(column=11+(k+1)*5, row=72).value = int(re.sub("[^0-9-,]","", line)[1:])
                ws.cell(column=12+(k+1)*5, row=72).value = int(re.sub("[^0-9-,]","", line)[1:])
            if "AWB_TAG_EXTRACOLOR_"+str(k)+"_GAIN_R" in line:
                ws.cell(column=9+(k+1)*5, row=74).value = int(re.sub("[^0-9-,]","", line)[1:])
                ws.cell(column=12+(k+1)*5, row=73).value = int(re.sub("[^0-9-,]","", line)[1:])
            if "AWB_TAG_EXTRACOLOR_"+str(k)+"_GAIN_B" in line:
                ws.cell(column=11+(k+1)*5, row=74).value = int(re.sub("[^0-9-,]","", line)[1:])
                ws.cell(column=12+(k+1)*5, row=74).value = int(re.sub("[^0-9-,]","", line)[1:])
            if "AWB_TAG_EXTRACOLOR_"+str(k)+"_GAVG_L" in line:
                ws.cell(column=9+(k+1)*5, row=77).value = int(re.sub("[^0-9-,]","", line)[1:])
            if "AWB_TAG_EXTRACOLOR_"+str(k)+"_GAVG_H" in line:
                ws.cell(column=9+(k+1)*5, row=78).value = int(re.sub("[^0-9-,]","", line)[1:])
            if "AWB_TAG_EXTRACOLOR_"+str(k)+"_LV_L" in line:
                ws.cell(column=11+(k+1)*5, row=77).value = int(re.sub("[^0-9-,]","", line)[1:])
                ws.cell(column=12+(k+1)*5, row=77).value = int(re.sub("[^0-9-,]","", line)[1:])
            if "AWB_TAG_EXTRACOLOR_"+str(k)+"_LV_H" in line:
                ws.cell(column=11+(k+1)*5, row=78).value = int(re.sub("[^0-9-,]","", line)[1:])
                ws.cell(column=12+(k+1)*5, row=78).value = int(re.sub("[^0-9-,]","", line)[1:])
            if "AWB_TAG_EXTRACOLOR_"+str(k)+"_COUNT_L" in line:
                ws.cell(column=9+(k+1)*5, row=79).value = int(re.sub("[^0-9-,]","", line)[1:])
                ws.cell(column=12+(k+1)*5, row=75).value = int(re.sub("[^0-9-,]","", line)[1:])
            if "AWB_TAG_EXTRACOLOR_"+str(k)+"_COUNT_H" in line:
                ws.cell(column=9+(k+1)*5, row=80).value = int(re.sub("[^0-9-,]","", line)[1:])
                ws.cell(column=12+(k+1)*5, row=76).value = int(re.sub("[^0-9-,]","", line)[1:])
            if "AWB_TAG_EXTRACOLOR_"+str(k)+"_WEI_L" in line:
                ws.cell(column=11+(k+1)*5, row=79).value = int(re.sub("[^0-9-,]","", line)[1:])
                ws.cell(column=12+(k+1)*5, row=79).value = int(re.sub("[^0-9-,]","", line)[1:])
            if "AWB_TAG_EXTRACOLOR_"+str(k)+"_WEI_H" in line:
                ws.cell(column=11+(k+1)*5, row=80).value = int(re.sub("[^0-9-,]","", line)[1:])
                ws.cell(column=12+(k+1)*5, row=80).value = int(re.sub("[^0-9-,]","", line)[1:])
        
        if "AWB_TAG_REFINE_P0_PB_RATIO" in line:
            ws.cell(column=12, row=32).value = int(re.sub("[^0-9-,]","", line)[1:])
        if "AWB_TAG_REFINE_P0_RANGE" in line:
            ws.cell(column=12, row=34).value = int(re.sub("[^0-9-,]","", line)[1:])
        if "AWB_TAG_REFINE_P0_CLIP" in line:
            ws.cell(column=12, row=36).value = int(re.sub("[^0-9-,]","", line)[1:])
        
        if "AWB_TAG_REFINE_P2_PB_RATIO" in line:
            ws.cell(column=18, row=57).value = int(re.sub("[^0-9-,]","", line)[1:])
        if "AWB_TAG_REFINE_P2_LV_THR" in line:
            ws.cell(column=19, row=57).value = int(re.sub("[^0-9-,]","", line)[1:])
        if "AWB_TAG_REFINE_P2_P0_THR1" in line:
            ws.cell(column=20, row=57).value = int(re.sub("[^0-9-,]","", line)[3:])
        if "AWB_TAG_REFINE_P2_P0_THR2" in line:
            ws.cell(column=21, row=57).value = int(re.sub("[^0-9-,]","", line)[3:])
        
        if "AWB_TAG_STA_GAIN_R_T" in line:
            ws.cell(column=16, row=31).value = int(re.sub("[^0-9-,]","", line))
        if "AWB_TAG_STA_GAIN_B_T" in line:
            ws.cell(column=16, row=32).value = int(re.sub("[^0-9-,]","", line))
        if "AWB_TAG_SPAT_GAIN_R_T" in line:
            ws.cell(column=16, row=33).value = int(re.sub("[^0-9-,]","", line))
        if "AWB_TAG_SPAT_GAIN_B_T" in line:
            ws.cell(column=16, row=34).value = int(re.sub("[^0-9-,]","", line))
        if "AWB_TAG_STA_GAIN_R_WF" in line:
            ws.cell(column=21, row=31).value = int(re.sub("[^0-9-,]","", line))
        if "AWB_TAG_STA_GAIN_B_WF" in line:
            ws.cell(column=21, row=32).value = int(re.sub("[^0-9-,]","", line))
        if "AWB_TAG_SPAT_GAIN_R_WF" in line:
            ws.cell(column=21, row=33).value = int(re.sub("[^0-9-,]","", line))
        if "AWB_TAG_SPAT_GAIN_B_WF" in line:
            ws.cell(column=21, row=34).value = int(re.sub("[^0-9-,]","", line))
        if "AWB_TAG_STA_GAIN_R_F" in line:
            ws.cell(column=26, row=31).value = int(re.sub("[^0-9-,]","", line))
        if "AWB_TAG_STA_GAIN_B_F" in line:
            ws.cell(column=26, row=32).value = int(re.sub("[^0-9-,]","", line))
        if "AWB_TAG_SPAT_GAIN_R_F" in line:
            ws.cell(column=26, row=33).value = int(re.sub("[^0-9-,]","", line))
        if "AWB_TAG_SPAT_GAIN_B_F" in line:
            ws.cell(column=26, row=34).value = int(re.sub("[^0-9-,]","", line))
        if "AWB_TAG_STA_GAIN_R_CWF" in line:
            ws.cell(column=31, row=31).value = int(re.sub("[^0-9-,]","", line))
        if "AWB_TAG_STA_GAIN_B_CWF" in line:
            ws.cell(column=31, row=32).value = int(re.sub("[^0-9-,]","", line))
        if "AWB_TAG_SPAT_GAIN_R_CWF" in line:
            ws.cell(column=31, row=33).value = int(re.sub("[^0-9-,]","", line))
        if "AWB_TAG_SPAT_GAIN_B_CWF" in line:
            ws.cell(column=31, row=34).value = int(re.sub("[^0-9-,]","", line))
        if "AWB_TAG_STA_GAIN_R_D" in line:
            AWB_TAG_STA_GAIN_R_D.append(re.sub("[^0-9-,]","", line))
        if "AWB_TAG_STA_GAIN_B_D" in line:
            AWB_TAG_STA_GAIN_B_D.append(re.sub("[^0-9-,]","", line))
        if "AWB_TAG_SPAT_GAIN_R_D" in line:
            AWB_TAG_SPAT_GAIN_R_D.append(re.sub("[^0-9-,]","", line))
        if "AWB_TAG_SPAT_GAIN_B_D" in line:
            AWB_TAG_SPAT_GAIN_B_D.append(re.sub("[^0-9-,]","", line))
        if "AWB_TAG_STA_GAIN_R_S" in line:
            ws.cell(column=41, row=31).value = int(re.sub("[^0-9-,]","", line))
        if "AWB_TAG_STA_GAIN_B_S" in line:
            ws.cell(column=41, row=32).value = int(re.sub("[^0-9-,]","", line))
        if "AWB_TAG_SPAT_GAIN_R_S" in line:
            ws.cell(column=41, row=33).value = int(re.sub("[^0-9-,]","", line))
        if "AWB_TAG_SPAT_GAIN_B_S" in line:
            ws.cell(column=41, row=34).value = int(re.sub("[^0-9-,]","", line))
        if "AWB_TAG_STA_GAIN_R_DF" in line:
            ws.cell(column=46, row=31).value = int(re.sub("[^0-9-,]","", line))
        if "AWB_TAG_STA_GAIN_B_DF" in line:
            ws.cell(column=46, row=32).value = int(re.sub("[^0-9-,]","", line))
        if "AWB_TAG_SPAT_GAIN_R_DF" in line:
            ws.cell(column=46, row=33).value = int(re.sub("[^0-9-,]","", line))
        if "AWB_TAG_SPAT_GAIN_B_DF" in line:
            ws.cell(column=46, row=34).value = int(re.sub("[^0-9-,]","", line))
            
        if "AWB_TAG_PREFGAIN_TUNGSTEN_R" in line:
            ws.cell(column=14, row=39).value = int(re.sub("[^0-9-,]","", line))
            ws.cell(column=1, row=39).value = int(re.sub("[^0-9-,]","", line))
        if "AWB_TAG_PREFGAIN_TUNGSTEN_B" in line:
            ws.cell(column=16, row=39).value = int(re.sub("[^0-9-,]","", line))
            ws.cell(column=17, row=39).value = int(re.sub("[^0-9-,]","", line))
        if "AWB_TAG_PREFGAIN_WF_R" in line:
            ws.cell(column=14, row=40).value = int(re.sub("[^0-9-,]","", line))
            ws.cell(column=1, row=40).value = int(re.sub("[^0-9-,]","", line))
        if "AWB_TAG_PREFGAIN_WF_B" in line:
            ws.cell(column=16, row=40).value = int(re.sub("[^0-9-,]","", line))
            ws.cell(column=17, row=40).value = int(re.sub("[^0-9-,]","", line))
        if "AWB_TAG_PREFGAIN_FLUORESCENT_R" in line:
            ws.cell(column=14, row=41).value = int(re.sub("[^0-9-,]","", line))
            ws.cell(column=1, row=41).value = int(re.sub("[^0-9-,]","", line))
        if "AWB_TAG_PREFGAIN_FLUORESCENT_B" in line:
            ws.cell(column=16, row=41).value = int(re.sub("[^0-9-,]","", line))
            ws.cell(column=17, row=41).value = int(re.sub("[^0-9-,]","", line))
        if "AWB_TAG_PREFGAIN_CWF_R" in line:
            ws.cell(column=14, row=42).value = int(re.sub("[^0-9-,]","", line))
            ws.cell(column=1, row=42).value = int(re.sub("[^0-9-,]","", line))
        if "AWB_TAG_PREFGAIN_CWF_B" in line:
            ws.cell(column=16, row=42).value = int(re.sub("[^0-9-,]","", line))
            ws.cell(column=17, row=42).value = int(re.sub("[^0-9-,]","", line))
        if "AWB_TAG_PREFGAIN_DAYLIGHT_R" in line:
            ws.cell(column=14, row=43).value = int(re.sub("[^0-9-,]","", line))
            ws.cell(column=1, row=43).value = int(re.sub("[^0-9-,]","", line))
        if "AWB_TAG_PREFGAIN_DAYLIGHT_B" in line:
            ws.cell(column=16, row=43).value = int(re.sub("[^0-9-,]","", line))
            ws.cell(column=17, row=43).value = int(re.sub("[^0-9-,]","", line))
        if "AWB_TAG_PREFGAIN_SHADE_R" in line:
            ws.cell(column=14, row=44).value = int(re.sub("[^0-9-,]","", line))
            ws.cell(column=1, row=44).value = int(re.sub("[^0-9-,]","", line))
        if "AWB_TAG_PREFGAIN_SHADE_B" in line:
            ws.cell(column=16, row=44).value = int(re.sub("[^0-9-,]","", line))
            ws.cell(column=17, row=44).value = int(re.sub("[^0-9-,]","", line))
        if "AWB_TAG_PREFGAIN_DF_R" in line:
            ws.cell(column=14, row=45).value = int(re.sub("[^0-9-,]","", line))
            ws.cell(column=1, row=45).value = int(re.sub("[^0-9-,]","", line))
        if "AWB_TAG_PREFGAIN_DF_B" in line:
            ws.cell(column=16, row=45).value = int(re.sub("[^0-9-,]","", line))
            ws.cell(column=17, row=45).value = int(re.sub("[^0-9-,]","", line))
        
        if "AWB_TAG_DAYLIGHT_PROB_T" in line:
            ws.cell(column=26, row=39).value = int(re.sub("[^0-9-,]","", line))
            ws.cell(column=27, row=39).value = int(re.sub("[^0-9-,]","", line))
        if "AWB_TAG_DAYLIGHT_PROB_WF" in line:
            ws.cell(column=26, row=40).value = int(re.sub("[^0-9-,]","", line))
            ws.cell(column=27, row=40).value = int(re.sub("[^0-9-,]","", line))
        if "AWB_TAG_DAYLIGHT_PROB_F" in line:
            ws.cell(column=26, row=41).value = int(re.sub("[^0-9-,]","", line))
            ws.cell(column=27, row=41).value = int(re.sub("[^0-9-,]","", line))
        if "AWB_TAG_DAYLIGHT_PROB_CWF" in line:
            ws.cell(column=26, row=42).value = int(re.sub("[^0-9-,]","", line))
            ws.cell(column=27, row=42).value = int(re.sub("[^0-9-,]","", line))
        if "AWB_TAG_DAYLIGHT_PROB_D" in line:
            AWB_TAG_DAYLIGHT_PROB_D.append(re.sub("[^0-9-,]","", line))
        if "AWB_TAG_DAYLIGHT_PROB_S" in line:
            ws.cell(column=26, row=44).value = int(re.sub("[^0-9-,]","", line))
            ws.cell(column=27, row=44).value = int(re.sub("[^0-9-,]","", line))
        if "AWB_TAG_DAYLIGHT_PROB_DF" in line:
            ws.cell(column=26, row=45).value = int(re.sub("[^0-9-,]","", line))
            ws.cell(column=27, row=45).value = int(re.sub("[^0-9-,]","", line))
            
        if "AWB_NVRAM_SHADE_F_DAYLIGHT_PROB" in line:
            ws.cell(column=31, row=40).value = int(re.sub("[^0-9-,]","", line))
            ws.cell(column=32, row=40).value = int(re.sub("[^0-9-,]","", line))
        if "AWB_NVRAM_SHADE_F_AREA_XR_THR" in line:
            Sub_F_area.append(re.sub("[^0-9-,]","", line))
            ws.cell(column=29, row=41).value = int(re.sub("[^0-9-,]","", line))
        if "AWB_NVRAM_SHADE_F_AREA_YR_THR" in line:
            Sub_F_area.append(re.sub("[^0-9-,]","", line))
            ws.cell(column=31, row=41).value = int(re.sub("[^0-9-,]","", line))
        if "AWB_NVRAM_FLUORESCENT_RIGHT" in line:
            Sub_F_area.append(re.sub("[^0-9-,]","", line))
            ws.cell(column=29, row=42).value = int(re.sub("[^0-9-,]","", line))
        if "AWB_NVRAM_FLUORESCENT_LOWER" in line:
            Sub_F_area.append(re.sub("[^0-9-,]","", line))
            ws.cell(column=31, row=42).value = int(re.sub("[^0-9-,]","", line))
        if "AWB_NVRAM_SHADE_F_VERTEX_XR_THR" in line:
            ws.cell(column=29, row=43).value = int(re.sub("[^0-9-,]","", line))
            ws.cell(column=32, row=44).value = int(re.sub("[^0-9-,]","", line))
        if "AWB_NVRAM_SHADE_F_VERTEX_YR_THR" in line:
            ws.cell(column=31, row=43).value = int(re.sub("[^0-9-,]","", line))
            ws.cell(column=32, row=45).value = int(re.sub("[^0-9-,]","", line))
        if "AWB_TAG_SHADE_XR_F" in line:
            ws.cell(column=29, row=44).value = int(re.sub("[^0-9-,]","", line))
        if "AWB_TAG_SHADE_YR_F" in line:
            ws.cell(column=31, row=44).value = int(re.sub("[^0-9-,]","", line))
        if "AWB_NVRAM_SHADE_F_LV_THR_L" in line:
            ws.cell(column=34, row=40).value = int(re.sub("[^0-9-,]","", line))
            ws.cell(column=32, row=41).value = int(re.sub("[^0-9-,]","", line))
        if "AWB_NVRAM_SHADE_F_LV_THR_H" in line:
            ws.cell(column=36, row=40).value = int(re.sub("[^0-9-,]","", line))
            ws.cell(column=32, row=42).value = int(re.sub("[^0-9-,]","", line))
        if "AWB_TAG_SHADE_COUNT_F" in line:
            ws.cell(column=36, row=42).value = int(re.sub("[^0-9-,]","", line))
            ws.cell(column=32, row=43).value = int(re.sub("[^0-9-,]","", line))
            
        if "AWB_NVRAM_SHADE_CWF_DAYLIGHT_PROB" in line:
            ws.cell(column=41, row=40).value = int(re.sub("[^0-9-,]","", line))
            ws.cell(column=42, row=40).value = int(re.sub("[^0-9-,]","", line))
        if "AWB_NVRAM_SHADE_CWF_AREA_XR_THR" in line:
            Sub_CWF_area.append(re.sub("[^0-9-,]","", line))
            ws.cell(column=39, row=41).value = int(re.sub("[^0-9-,]","", line))
        if "AWB_NVRAM_SHADE_CWF_AREA_YR_THR" in line:
            Sub_CWF_area.append(re.sub("[^0-9-,]","", line))
            ws.cell(column=41, row=41).value = int(re.sub("[^0-9-,]","", line))
        if "AWB_NVRAM_CWF_RIGHT" in line:
            Sub_CWF_area.append(re.sub("[^0-9-,]","", line))
            ws.cell(column=39, row=42).value = int(re.sub("[^0-9-,]","", line))
        if "AWB_NVRAM_CWF_LOWER" in line:
            Sub_CWF_area.append(re.sub("[^0-9-,]","", line))
            ws.cell(column=41, row=42).value = int(re.sub("[^0-9-,]","", line))
        if "AWB_NVRAM_SHADE_CWF_VERTEX_XR_THR" in line:
            ws.cell(column=39, row=43).value = int(re.sub("[^0-9-,]","", line))
            ws.cell(column=42, row=44).value = int(re.sub("[^0-9-,]","", line))
        if "AWB_NVRAM_SHADE_CWF_VERTEX_YR_THR" in line:
            ws.cell(column=41, row=43).value = int(re.sub("[^0-9-,]","", line))
            ws.cell(column=42, row=45).value = int(re.sub("[^0-9-,]","", line))
        if "AWB_TAG_SHADE_XR_CWF" in line:
            ws.cell(column=39, row=44).value = int(re.sub("[^0-9-,]","", line))
        if "AWB_TAG_SHADE_YR_CWF" in line:
            ws.cell(column=41, row=44).value = int(re.sub("[^0-9-,]","", line))
        if "AWB_NVRAM_SHADE_CWF_LV_THR_L" in line:
            ws.cell(column=44, row=40).value = int(re.sub("[^0-9-,]","", line))
            ws.cell(column=42, row=41).value = int(re.sub("[^0-9-,]","", line))
        if "AWB_NVRAM_SHADE_CWF_LV_THR_H" in line:
            ws.cell(column=46, row=40).value = int(re.sub("[^0-9-,]","", line))
            ws.cell(column=42, row=42).value = int(re.sub("[^0-9-,]","", line))
        if "AWB_TAG_SHADE_COUNT_CWF" in line:
            ws.cell(column=46, row=42).value = int(re.sub("[^0-9-,]","", line))
            ws.cell(column=42, row=43).value = int(re.sub("[^0-9-,]","", line))
        
        if "AWB_TAG_STAT_LIMIT_LV_L" in line:
            ws.cell(column=12, row=51).value = int(re.sub("[^0-9-,]","", line))
        if "AWB_TAG_STAT_LIMIT_LV_H" in line:
            ws.cell(column=12, row=54).value = int(re.sub("[^0-9-,]","", line))
        if "AWB_NVRAM_D65_YR" in line:
            ws.cell(column=16, row=49).value = int(re.sub("[^0-9-,]","", line)[2:])
            ws.cell(column=21, row=49).value = int(re.sub("[^0-9-,]","", line)[2:])
            ws.cell(column=26, row=49).value = int(re.sub("[^0-9-,]","", line)[2:])
            ws.cell(column=31, row=49).value = int(re.sub("[^0-9-,]","", line)[2:])
            ws.cell(column=36, row=49).value = int(re.sub("[^0-9-,]","", line)[2:])
            ws.cell(column=41, row=49).value = int(re.sub("[^0-9-,]","", line)[2:])
            ws.cell(column=46, row=49).value = int(re.sub("[^0-9-,]","", line)[2:])
        if "AWB_TAG_AVG_YR_T" in line:
            ws.cell(column=16, row=50).value = int(re.sub("[^0-9-,]","", line))
        if "AWB_TAG_AVG_YR_WF" in line:
            ws.cell(column=21, row=50).value = int(re.sub("[^0-9-,]","", line))
        if "AWB_TAG_AVG_YR_F" in line:
            ws.cell(column=26, row=50).value = int(re.sub("[^0-9-,]","", line))
        if "AWB_TAG_AVG_YR_CWF" in line:
            ws.cell(column=31, row=50).value = int(re.sub("[^0-9-,]","", line))
        if "AWB_TAG_AVG_YR_D" in line:
            AWB_TAG_AVG_YR_D.append(re.sub("[^0-9-,]","", line))
        if "AWB_TAG_AVG_YR_S" in line:
            ws.cell(column=41, row=50).value = int(re.sub("[^0-9-,]","", line))
        if "AWB_TAG_AVG_YR_DF" in line:
            ws.cell(column=46, row=50).value = int(re.sub("[^0-9-,]","", line))
        if "AWB_TAG_STAT_LIMIT_Y_T" in line:
            ws.cell(column=14, row=51).value = int(re.sub("[^0-9-,]","", line))
        if "AWB_TAG_STAT_LIMIT_Y_WF" in line:
            ws.cell(column=19, row=51).value = int(re.sub("[^0-9-,]","", line))
        if "AWB_TAG_STAT_LIMIT_Y_F" in line:
            ws.cell(column=24, row=51).value = int(re.sub("[^0-9-,]","", line))
        if "AWB_TAG_STAT_LIMIT_Y_CWF" in line:
            ws.cell(column=29, row=51).value = int(re.sub("[^0-9-,]","", line))
        if "AWB_TAG_STAT_LIMIT_Y_D" in line:
            AWB_TAG_STAT_LIMIT_Y_D.append(re.sub("[^0-9-,]","", line))
        if "AWB_TAG_STAT_LIMIT_Y_S" in line:
            ws.cell(column=39, row=51).value = int(re.sub("[^0-9-,]","", line))
        if "AWB_TAG_STAT_LIMIT_Y_DF" in line:
            ws.cell(column=44, row=51).value = int(re.sub("[^0-9-,]","", line))
        if "AWB_TAG_LIMIT_T" in line:
            ws.cell(column=16, row=51).value = int(re.sub("[^0-9-,]","", line))
            ws.cell(column=17, row=51).value = int(re.sub("[^0-9-,]","", line))
        if "AWB_TAG_LIMIT_WF" in line:
            ws.cell(column=21, row=51).value = int(re.sub("[^0-9-,]","", line))
            ws.cell(column=22, row=51).value = int(re.sub("[^0-9-,]","", line))
        if "AWB_TAG_LIMIT_F" in line:
            ws.cell(column=26, row=51).value = int(re.sub("[^0-9-,]","", line))
            ws.cell(column=27, row=51).value = int(re.sub("[^0-9-,]","", line))
        if "AWB_TAG_LIMIT_CWF" in line:
            ws.cell(column=31, row=51).value = int(re.sub("[^0-9-,]","", line))
            ws.cell(column=32, row=51).value = int(re.sub("[^0-9-,]","", line))
        if "AWB_TAG_LIMIT_D" in line:
            AWB_TAG_LIMIT_D.append(re.sub("[^0-9-,]","", line))
        if "AWB_TAG_LIMIT_S" in line:
            ws.cell(column=41, row=51).value = int(re.sub("[^0-9-,]","", line))
            ws.cell(column=42, row=51).value = int(re.sub("[^0-9-,]","", line))
        if "AWB_TAG_LIMIT_DF" in line:
            ws.cell(column=46, row=51).value = int(re.sub("[^0-9-,]","", line))
            ws.cell(column=47, row=51).value = int(re.sub("[^0-9-,]","", line))
        if "AWB_TAG_STAT_LIMIT_W_RED_T" in line:
            ws.cell(column=14, row=53).value = int(re.sub("[^0-9-,]","", line))
        if "AWB_TAG_STAT_LIMIT_W_RED_WF" in line:
            ws.cell(column=19, row=53).value = int(re.sub("[^0-9-,]","", line))
        if "AWB_TAG_STAT_LIMIT_W_RED_F" in line:
            ws.cell(column=24, row=53).value = int(re.sub("[^0-9-,]","", line))
        if "AWB_TAG_STAT_LIMIT_W_RED_CWF" in line:
            ws.cell(column=29, row=53).value = int(re.sub("[^0-9-,]","", line))
        if "AWB_TAG_STAT_LIMIT_W_RED_D" in line:
            AWB_TAG_STAT_LIMIT_W_RED_D.append(re.sub("[^0-9-,]","", line))
        if "AWB_TAG_STAT_LIMIT_W_RED_S" in line:
            ws.cell(column=39, row=53).value = int(re.sub("[^0-9-,]","", line))
        if "AWB_TAG_STAT_LIMIT_W_RED_DF" in line:
            ws.cell(column=44, row=53).value = int(re.sub("[^0-9-,]","", line))
        if "AWB_TAG_STAT_LIMIT_PROJ_W_T" in line:
            ws.cell(column=16, row=53).value = int(re.sub("[^0-9-,]","", line))
        if "AWB_TAG_STAT_LIMIT_PROJ_W_WF" in line:
            ws.cell(column=21, row=53).value = int(re.sub("[^0-9-,]","", line))
        if "AWB_TAG_STAT_LIMIT_PROJ_W_F" in line:
            ws.cell(column=26, row=53).value = int(re.sub("[^0-9-,]","", line))
        if "AWB_TAG_STAT_LIMIT_PROJ_W_CWF" in line:
            ws.cell(column=31, row=53).value = int(re.sub("[^0-9-,]","", line))
        if "AWB_TAG_STAT_LIMIT_PROJ_W_D" in line:
            AWB_TAG_STAT_LIMIT_PROJ_W_D.append(re.sub("[^0-9-,]","", line))
        if "AWB_TAG_STAT_LIMIT_PROJ_W_S" in line:
            ws.cell(column=41, row=53).value = int(re.sub("[^0-9-,]","", line))
        if "AWB_TAG_STAT_LIMIT_PROJ_W_DF" in line:
            ws.cell(column=46, row=53).value = int(re.sub("[^0-9-,]","", line))
        
        if "AWB_TAG_PB_NUM_THR_T" in line:
            ws.cell(column=24, row=57).value = int(re.sub("[^0-9-,]","", line))
            ws.cell(column=22, row=57).value = int(re.sub("[^0-9-,]","", line))
        if "AWB_TAG_PB_NUM_THR_WF" in line:
            ws.cell(column=24, row=58).value = int(re.sub("[^0-9-,]","", line))
            ws.cell(column=22, row=58).value = int(re.sub("[^0-9-,]","", line))
        if "AWB_TAG_PB_NUM_THR_F" in line:
            ws.cell(column=24, row=59).value = int(re.sub("[^0-9-,]","", line))
            ws.cell(column=22, row=59).value = int(re.sub("[^0-9-,]","", line))
        if "AWB_TAG_PB_NUM_THR_CWF" in line:
            ws.cell(column=24, row=60).value = int(re.sub("[^0-9-,]","", line))
            ws.cell(column=22, row=60).value = int(re.sub("[^0-9-,]","", line))
        if "AWB_TAG_PB_NUM_THR_D" in line:
            AWB_TAG_PB_NUM_THR_D.append(re.sub("[^0-9-,]","", line))
        if "AWB_TAG_PB_NUM_THR_S" in line:
            ws.cell(column=24, row=62).value = int(re.sub("[^0-9-,]","", line))
            ws.cell(column=22, row=62).value = int(re.sub("[^0-9-,]","", line))
        if "AWB_TAG_PB_NUM_THR_DF" in line:
            ws.cell(column=24, row=63).value = int(re.sub("[^0-9-,]","", line))
            ws.cell(column=22, row=63).value = int(re.sub("[^0-9-,]","", line))
        # if "AWB_TAG_HIT_NR_T" in line:
        #     ws.cell(column=26, row=57).value = int(re.sub("[^0-9-,]","", line))
        # if "AWB_TAG_HIT_NR_WF" in line:
        #     ws.cell(column=26, row=58).value = int(re.sub("[^0-9-,]","", line))
        # if "AWB_TAG_HIT_NR_F" in line:
        #     ws.cell(column=26, row=59).value = int(re.sub("[^0-9-,]","", line))
        # if "AWB_TAG_HIT_NR_CWF" in line:
        #     ws.cell(column=26, row=60).value = int(re.sub("[^0-9-,]","", line))
        # if "AWB_TAG_HIT_NR_D" in line:
        #     AWB_TAG_HIT_NR_D.append(re.sub("[^0-9-,]","", line))
        # if "AWB_TAG_HIT_NR_S" in line:
        #     ws.cell(column=26, row=62).value = int(re.sub("[^0-9-,]","", line))
        # if "AWB_TAG_HIT_NR_DF" in line:
        #     ws.cell(column=26, row=63).value = int(re.sub("[^0-9-,]","", line))
            
        if "AWB_TAG_SPAT_LV_THR_L_T" in line:
            ws.cell(column=20, row=11).value = int(re.sub("[^0-9-,]","", line))
        if "AWB_TAG_SPAT_LV_THR_H_T" in line:
            ws.cell(column=21, row=11).value = int(re.sub("[^0-9-,]","", line))
        if "AWB_TAG_SPAT_LV_THR_L_WF" in line:
            ws.cell(column=20, row=12).value = int(re.sub("[^0-9-,]","", line))
        if "AWB_TAG_SPAT_LV_THR_H_WF" in line:
            ws.cell(column=21, row=12).value = int(re.sub("[^0-9-,]","", line))
        if "AWB_TAG_SPAT_LV_THR_L_F" in line:
            ws.cell(column=20, row=13).value = int(re.sub("[^0-9-,]","", line))
        if "AWB_TAG_SPAT_LV_THR_H_F" in line:
            ws.cell(column=21, row=13).value = int(re.sub("[^0-9-,]","", line))
        if "AWB_TAG_SPAT_LV_THR_L_CWF" in line:
            ws.cell(column=20, row=14).value = int(re.sub("[^0-9-,]","", line))
        if "AWB_TAG_SPAT_LV_THR_H_CWF" in line:
            ws.cell(column=21, row=14).value = int(re.sub("[^0-9-,]","", line))
        if "AWB_TAG_SPAT_LV_THR_L_D" in line:
            AWB_TAG_SPAT_LV_THR_L_D.append(re.sub("[^0-9-,]","", line))
        if "AWB_TAG_SPAT_LV_THR_H_D" in line:
            AWB_TAG_SPAT_LV_THR_H_D.append(re.sub("[^0-9-,]","", line))
        if "AWB_TAG_SPAT_LV_THR_L_S" in line:
            ws.cell(column=20, row=16).value = int(re.sub("[^0-9-,]","", line))
        if "AWB_TAG_SPAT_LV_THR_H_S" in line:
            ws.cell(column=21, row=16).value = int(re.sub("[^0-9-,]","", line))
        if "AWB_TAG_SPAT_LV_THR_L_DF" in line:
            ws.cell(column=20, row=17).value = int(re.sub("[^0-9-,]","", line))
        if "AWB_TAG_SPAT_LV_THR_H_DF" in line:
            ws.cell(column=21, row=17).value = int(re.sub("[^0-9-,]","", line))
        if "AWB_NVRAM_PREDICTOR_INIT_LV_THR_L" in line:
            ws.cell(column=18, row=15).value = int(re.sub("[^0-9-,]","", line))
        if "AWB_NVRAM_PREDICTOR_INIT_LV_THR_H" in line:
            ws.cell(column=18, row=17).value = int(re.sub("[^0-9-,]","", line))
        
        if "AWB_NVRAM_TUNGSTEN_RIGHT" in line:
            T_area.append(re.sub("[^0-9-,]","", line))
        if "AWB_NVRAM_TUNGSTEN_LEFT" in line:
            T_area.append(re.sub("[^0-9-,]","", line))
        if "AWB_NVRAM_TUNGSTEN_UPPER" in line:
            T_area.append(re.sub("[^0-9-,]","", line))
        if "AWB_NVRAM_TUNGSTEN_LOWER" in line:
            T_area.append(re.sub("[^0-9-,]","", line))
            
        if "AWB_NVRAM_WF_RIGHT" in line:
            WF_area.append(re.sub("[^0-9-,]","", line))
        if "AWB_NVRAM_WF_LEFT" in line:
            WF_area.append(re.sub("[^0-9-,]","", line))
        if "AWB_NVRAM_WF_UPPER" in line:
            WF_area.append(re.sub("[^0-9-,]","", line))
        if "AWB_NVRAM_WF_LOWER" in line:
            WF_area.append(re.sub("[^0-9-,]","", line))
            
        if "AWB_NVRAM_FLUORESCENT_RIGHT" in line:
            F_area.append(re.sub("[^0-9-,]","", line))
        if "AWB_NVRAM_FLUORESCENT_LEFT" in line:
            F_area.append(re.sub("[^0-9-,]","", line))
        if "AWB_NVRAM_FLUORESCENT_UPPER" in line:
            F_area.append(re.sub("[^0-9-,]","", line))
        if "AWB_NVRAM_FLUORESCENT_LOWER" in line:
            F_area.append(re.sub("[^0-9-,]","", line))
            
        if "AWB_NVRAM_CWF_RIGHT" in line:
            CWF_area.append(re.sub("[^0-9-,]","", line))
        if "AWB_NVRAM_CWF_LEFT" in line:
            CWF_area.append(re.sub("[^0-9-,]","", line))
        if "AWB_NVRAM_CWF_UPPER" in line:
            CWF_area.append(re.sub("[^0-9-,]","", line))
        if "AWB_NVRAM_CWF_LOWER" in line:
            CWF_area.append(re.sub("[^0-9-,]","", line))
            
        if "AWB_NVRAM_DAYLIGHT_RIGHT" in line:
            D_area.append(re.sub("[^0-9-,]","", line))
        if "AWB_NVRAM_DAYLIGHT_LEFT" in line:
            D_area.append(re.sub("[^0-9-,]","", line))
        if "AWB_NVRAM_DAYLIGHT_UPPER" in line:
            D_area.append(re.sub("[^0-9-,]","", line))
        if "AWB_NVRAM_DAYLIGHT_LOWER" in line:
            D_area.append(re.sub("[^0-9-,]","", line))
            
        if "AWB_NVRAM_SHADE_RIGHT" in line:
            S_area.append(re.sub("[^0-9-,]","", line))
        if "AWB_NVRAM_SHADE_LEFT" in line:
            S_area.append(re.sub("[^0-9-,]","", line))
        if "AWB_NVRAM_SHADE_UPPER" in line:
            S_area.append(re.sub("[^0-9-,]","", line))
        if "AWB_NVRAM_SHADE_LOWER" in line:
            S_area.append(re.sub("[^0-9-,]","", line))
            
        if "AWB_NVRAM_DF_RIGHT" in line:
            DF_area.append(re.sub("[^0-9-,]","", line))
        if "AWB_NVRAM_DF_LEFT" in line:
            DF_area.append(re.sub("[^0-9-,]","", line))
        if "AWB_NVRAM_DF_UPPER" in line:
            DF_area.append(re.sub("[^0-9-,]","", line))
        if "AWB_NVRAM_DF_LOWER" in line:
            DF_area.append(re.sub("[^0-9-,]","", line))
        
    ws.cell(column=10, row=4).value = int(CCT[0])
    ws.cell(column=5, row=24).value = int(AWB_TAG_NEUTRAL_PB_NUM_D[0])
    ws.cell(column=5, row=61).value = int(AWB_TAG_P1_D[0])
    ws.cell(column=1, row=61).value = int(AWB_TAG_P1_D[0])
    ws.cell(column=10, row=61).value = int(AWB_TAG_P2_D[0])
    ws.cell(column=36, row=31).value = int(AWB_TAG_STA_GAIN_R_D[0])
    ws.cell(column=36, row=32).value = int(AWB_TAG_STA_GAIN_B_D[0])
    ws.cell(column=36, row=33).value = int(AWB_TAG_SPAT_GAIN_R_D[0])
    ws.cell(column=36, row=34).value = int(AWB_TAG_SPAT_GAIN_B_D[0])
    ws.cell(column=26, row=43).value = int(AWB_TAG_DAYLIGHT_PROB_D[0])
    ws.cell(column=27, row=43).value = int(AWB_TAG_DAYLIGHT_PROB_D[0])
    ws.cell(column=36, row=50).value = int(AWB_TAG_AVG_YR_D[0])
    ws.cell(column=34, row=51).value = int(AWB_TAG_STAT_LIMIT_Y_D[0])
    ws.cell(column=36, row=51).value = int(AWB_TAG_LIMIT_D[0])
    ws.cell(column=37, row=51).value = int(AWB_TAG_LIMIT_D[0])
    ws.cell(column=34, row=53).value = int(AWB_TAG_STAT_LIMIT_W_RED_D[0])
    ws.cell(column=36, row=53).value = int(AWB_TAG_STAT_LIMIT_PROJ_W_D[0])
    ws.cell(column=24, row=61).value = int(AWB_TAG_PB_NUM_THR_D[0])
    ws.cell(column=22, row=61).value = int(AWB_TAG_PB_NUM_THR_D[0])
    ws.cell(column=20, row=15).value = int(AWB_TAG_SPAT_LV_THR_L_D[0])
    ws.cell(column=21, row=15).value = int(AWB_TAG_SPAT_LV_THR_H_D[0])
    # ws.cell(column=26, row=61).value = int(AWB_TAG_HIT_NR_D[0])
    
    for j in range(0,(np.size(allFileList_jpg))):
        path_name_jpg = yourPath + "/" + allFileList_jpg[j]
        file_name_jpg = os.path.basename(path_name_jpg)
        base2 = os.path.splitext(file_name_jpg)[0]
        if file_name_jpg == base or base2 == base[0:-8]:
            img = cv2.imread(path_name_jpg)
            height, width = img.shape[0], img.shape[1]
            
            if refer == 1:
                if j % 2 == 0:
                    path_name_jpg2 = yourPath + "/" + allFileList_jpg[j+1]
                else:
                    path_name_jpg2 = yourPath + "/" + allFileList_jpg[j-1]
                
                file_name_jpg2 = os.path.basename(path_name_jpg2)
                img2 = cv2.imread(path_name_jpg2)
                height2, width2 = img2.shape[0], img2.shape[1]
                
                save_img2 = openpyxl.drawing.image.Image(path_name_jpg2)
                if height > width and save_img2.height < save_img2.width:
                    rotate_name = yourPath + "/" + os.path.splitext(file_name_jpg2)[0] + "_rotate.png"
                    img2_rotate = Image.open(path_name_jpg2)
                    img2_rotate = img2_rotate.rotate(270, expand = True)
                    img2_rotate.save(rotate_name)
                    save_img2 = openpyxl.drawing.image.Image(rotate_name)
                    save_img2.height = 176
                    save_img2.width = 176 * width2 / height2 
                    save_img2.anchor = 'Z11'
                    ws.add_image(save_img2)
                elif height > width and save_img2.height > save_img2.width:
                    save_img2.height = 176
                    save_img2.width = 176 * width2 / height2
                    save_img2.anchor = 'Z11'
                    ws.add_image(save_img2)
                elif height < width and save_img2.height > save_img2.width:
                    rotate_name = yourPath + "/" + os.path.splitext(file_name_jpg2)[0] + "_rotate.png"
                    img2_rotate = Image.open(path_name_jpg2)
                    img2_rotate = img2_rotate.rotate(270, expand = True)
                    img2_rotate.save(rotate_name)
                    save_img2 = openpyxl.drawing.image.Image(rotate_name)
                    save_img2.height = 176
                    save_img2.width = 176 * width2 / height2 
                    save_img2.anchor = 'Z11'
                    ws.add_image(save_img2)
                else:
                    save_img2.height = 176
                    save_img2.width = 176 * width2 / height2
                    save_img2.anchor = 'Z11'
                    ws.add_image(save_img2)
            
            if height > width:
                save_img = openpyxl.drawing.image.Image(path_name_jpg)
                save_img.height = 176
                save_img.width = 176 * width / height
                save_img.anchor = 'Z2'
                ws.add_image(save_img)
            else:
                save_img = openpyxl.drawing.image.Image(path_name_jpg)
                save_img.height = 176
                save_img.width = 176 * width / height
                save_img.anchor = 'Z2'
                ws.add_image(save_img)
                
            plt.figure(figsize=(8.97,2.35))
            save_name = yourPath + "/" + os.path.splitext(file_name_jpg)[0] + "_grayworld.png"
            plt.tick_params(axis='x', which='major', labelsize=6)
            plt.tick_params(axis='y', which='major', labelsize=6)
            plt.xlim(-950,750)
            plt.ylim(-950,150)
            ax = plt.gca()
            
            rect_T = patches.Rectangle((int(T_area[1]),int(T_area[3])),
                 int(T_area[0])-int(T_area[1]),
                 int(T_area[2])-int(T_area[3]),
                 linewidth=1.5,
                 edgecolor='red',
                 fill = False)
            ax.add_patch(rect_T)
            rect_WF = patches.Rectangle((int(WF_area[1]),int(WF_area[3])),
                 int(WF_area[0])-int(WF_area[1]),
                 int(WF_area[2])-int(WF_area[3]),
                 linewidth=1.5,
                 edgecolor='orange',
                 fill = False)
            ax.add_patch(rect_WF)
            rect_F = patches.Rectangle((int(F_area[1]),int(F_area[3])),
                 int(F_area[0])-int(F_area[1]),
                 int(F_area[2])-int(F_area[3]),
                 linewidth=1.5,
                 edgecolor='yellow',
                 fill = False)
            ax.add_patch(rect_F)
            rect_sub_F = patches.Rectangle((int(Sub_F_area[2]),int(Sub_F_area[1])),
                 int(Sub_F_area[0])-int(Sub_F_area[2]),
                 int(Sub_F_area[3])-int(Sub_F_area[1]),
                 linewidth=1.5,
                 edgecolor='khaki',
                 fill = False)
            ax.add_patch(rect_sub_F)
            rect_CWF = patches.Rectangle((int(CWF_area[1]),int(CWF_area[3])),
                 int(CWF_area[0])-int(CWF_area[1]),
                 int(CWF_area[2])-int(CWF_area[3]),
                 linewidth=1.5,
                 edgecolor='lime',
                 fill = False)
            ax.add_patch(rect_CWF)
            rect_sub_CWF = patches.Rectangle((int(Sub_CWF_area[2]),int(Sub_CWF_area[1])),
                 int(Sub_CWF_area[0])-int(Sub_CWF_area[2]),
                 int(Sub_CWF_area[3])-int(Sub_CWF_area[1]),
                 linewidth=1.5,
                 edgecolor='springgreen',
                 fill = False)
            ax.add_patch(rect_sub_CWF)
            rect_D = patches.Rectangle((int(D_area[1]),int(D_area[3])),
                 int(D_area[0])-int(D_area[1]),
                 int(D_area[2])-int(D_area[3]),
                 linewidth=1.5,
                 edgecolor='blue',
                 fill = False)
            ax.add_patch(rect_D)
            rect_S = patches.Rectangle((int(S_area[1]),int(S_area[3])),
                 int(S_area[0])-int(S_area[1]),
                 int(S_area[2])-int(S_area[3]),
                 linewidth=1.5,
                 edgecolor='darkblue',
                 fill = False)
            ax.add_patch(rect_S)
            rect_DF = patches.Rectangle((int(DF_area[1]),int(DF_area[3])),
                 int(DF_area[0])-int(DF_area[1]),
                 int(DF_area[2])-int(DF_area[3]),
                 linewidth=1.5,
                 edgecolor='cyan',
                 fill = False)
            ax.add_patch(rect_DF)
            
            for t in range(0,8):
                rect_exclude = patches.Rectangle((int(locals()['exclude_'+str(t)][2]),int(locals()['exclude_'+str(t)][1])),
                     int(locals()['exclude_'+str(t)][3])-int(locals()['exclude_'+str(t)][2]),
                     int(locals()['exclude_'+str(t)][0])-int(locals()['exclude_'+str(t)][1]),
                     linewidth=1,
                     edgecolor='whitesmoke',
                     fill = False)
                ax.add_patch(rect_exclude)
                
            for t in range(0,8):
                if locals()['AWB_TAG_EXTRACOLOR_ENABLE'][t] == 1:
                    rect_extra = patches.Rectangle((int(locals()['extra_'+str(t)][2]),int(locals()['extra_'+str(t)][1])),
                         int(locals()['extra_'+str(t)][3])-int(locals()['extra_'+str(t)][2]),
                         int(locals()['extra_'+str(t)][0])-int(locals()['extra_'+str(t)][1]),
                         linewidth=1,
                         edgecolor='whitesmoke',
                         fill = False)
                    ax.add_patch(rect_extra)
            
            ax.spines['left'].set_position('zero')
            ax.spines['bottom'].set_position('zero')
            ax.spines['top'].set_color('none')
            ax.spines['right'].set_color('none')
            ax.set_facecolor('lightgrey')
            plt.savefig(save_name,bbox_inches='tight',dpi=100,pad_inches=0)
            plt.close()
            save_img_hist = openpyxl.drawing.image.Image(save_name)
            save_img_hist.anchor = 'O2'
            ws.add_image(save_img_hist)
            
            if i % 20 == 19:
                endNum = re.sub("[^0-9-,]","", base[0:2])
                file = "mtkAWBanalysis_" + str(localtime[0]) + "_" + str(localtime[1]) + "_" + str(localtime[2]) + "_" + clock + "_" + startNum + "_" + endNum + ".xlsm"
                wb.active = 0
                wb.save(file)
            break

endNum = re.sub("[^0-9-,]","", base[0:2])
file = "mtkAWBanalysis_" + str(localtime[0]) + "_" + str(localtime[1]) + "_" + str(localtime[2]) + "_" + clock + "_" + startNum + "_" + endNum + ".xlsm"
wb.active = 0
wb.save(file)

print("mtkAWBanalysis is ok!")
os.system("pause")