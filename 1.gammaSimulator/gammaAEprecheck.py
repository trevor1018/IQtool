import openpyxl
import csv
import numpy as np
import time
import os

def file_filter(f):
    if f[-4:] in ['.csv']:
        return True
    else:
        return False

print("gammaAEprecheck is runing...")
localtime = time.localtime()
clock = str(60*60*localtime[3] + 60*localtime[4] + localtime[5])

fn = 'gammaSummary.xlsm'
wb = openpyxl.load_workbook(fn, read_only=False, keep_vba=True)
wb.active = 0
ws = wb.active

yourPath1 = "0.target"
allFileList1 = os.listdir(yourPath1)
allFileList1 = np.sort(allFileList1,axis=0)
allFileList1 = list(filter(file_filter, allFileList1))

yourPath2 = "1.original"
allFileList2 = os.listdir(yourPath2)
allFileList2 = np.sort(allFileList2,axis=0)
allFileList2 = list(filter(file_filter, allFileList2))

order = ["18","9","6","10","7","4","11","12","13","8","5","14","15","16","17"]

for i in range(0,np.size(allFileList1)):
    print(allFileList1[i])
    
    path_name1 = yourPath1 + "/" + allFileList1[i]
    path_name2 = yourPath2 + "/" + allFileList2[i]
    file1 = open(path_name1)
    file2 = open(path_name2)
    reader1 = csv.reader(file1)
    reader2 = csv.reader(file2)
    data_list1 = list(reader1)
    data_list2 = list(reader2)
    
    for j in range(0,20,1):
        if data_list1[0][1] == "3.7":
            ws.cell(column=int(order[i]), row=j+12).value = float(data_list1[8+j][1])
            ws.cell(column=int(order[i])+18, row=j+12).value = float(data_list2[8+j][1])
        else:
            ws.cell(column=int(order[i]), row=j+12).value = float(data_list1[11+j][1])
            ws.cell(column=int(order[i])+18, row=j+12).value = float(data_list2[11+j][1])

        
file = "gammaSummary_" + str(localtime[0]) + "_" + str(localtime[1]) + "_" + str(localtime[2]) + "_" + clock + ".xlsm"
wb.active = 0
wb.save(file)

print("gammaAEprecheck is ok!")
os.system("pause")